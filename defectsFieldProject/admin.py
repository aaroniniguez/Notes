from stapp.models import TestCase, TestCaseAuthor, Tag, Variable, TestModule, TestSet, TestScript, MiscFileLocation, TagManager, TestExecutionHistory, Dependency, ScriptMiscFile, DBFlag
from django.contrib import admin
from django import forms
from django.http import HttpResponse, HttpResponseRedirect, Http404
from django.contrib.auth.models import User, Group
from django.contrib.auth.admin import UserAdmin
from django.contrib.sites.models import Site
from django.core.exceptions import ObjectDoesNotExist
from django.forms.formsets import formset_factory
from django.forms.formsets import BaseFormSet
import reversion
from django.utils.translation import ugettext_lazy as _
from django.contrib.admin import SimpleListFilter
from django.contrib import messages
from django.utils.safestring import mark_safe
from django.conf.urls import patterns
from django.shortcuts import render_to_response, redirect
from django.template import RequestContext
from django.utils.decorators import method_decorator
from django.db import models, transaction, router
from django.views.decorators.csrf import csrf_protect
from django.contrib.admin.util import unquote, flatten_fieldsets, get_deleted_objects, model_format_dict
from django.contrib.admin import widgets, helpers
from django.utils.encoding import force_unicode
from django.utils.html import escape, escapejs
from django.forms.formsets import all_valid
import traceback
from django.contrib.admin.models import LogEntry
import re
from django.utils import timezone
from copy import deepcopy
from django.db import transaction
from stapp.statemanager import StateManager
from django.contrib.admin.templatetags.admin_modify import *
from django.contrib.admin.templatetags.admin_modify import submit_row as original_submit_row
import sys
reload(sys)
sys.setdefaultencoding('utf-8')

TC_ID = ''
CURR_VAR_ID = 0
CURR_WRONG_VAR_ID = 0
VAR_EMPTY = False
to_edit = []
to_exclude = ['name',]
require_author = True

csrf_protect_m = method_decorator(csrf_protect)


@register.inclusion_tag('admin/submit_line.html', takes_context=True)
def submit_row(context):
    ctx = original_submit_row(context)
    ctx.update({
        'show_save_and_add_another': context.get('show_save_and_add_another', ctx['show_save_and_add_another']),
        'show_save_and_continue': context.get('show_save_and_continue', ctx['show_save_and_continue'])
        })
    return ctx

#@transaction.commit_manually
#def revertToActiveVerion(modeladmin, request, queryset):
#    msg = 'revert completed'
#    for tc in queryset:
#        if tc.state != 3 and tc.state != 4:
#            continue
#        if tc.previous_version is not None and tc.previous_version.state == 0:
#            tcname = ''
#            try:
#                tcname = tc.name
#                tc.delete()
#            except:
#                transaction.rollback()
#                msg = 'Error occured when reverting %s . Transaction rolled back' % tcname
#                break
#    transaction.commit()
#    messages.info(request, msg)
#revertToActiveVerion.short_description = 'Revert to active version'

def mod_str(str):
    if str.strip() == '':
        return ''
    return '"'+str+'"'

def tc_to_csv(modeladmin, request, queryset):
    time_str = timezone.now().strftime('%Y-%m-%d_%H%M%S%f')
    with open('/home/SmartTest/WebINF/stsite/static/download/temp/csv_export_'+time_str+'.csv', 'w') as f:
        f.write('TestName,TestScriptName,Variables,Dependencies,ScriptMiscFiles,TestModule,Tags,Author,State,Procedure,P4ScriptVersion,ScriptType,Duration,Defects,TimeOut,MRQ,TestSet,TimeStamp,TestType,MinSTCVersion,ExecHarness,MST,Priority\r');
        try:
            for testcase in queryset:
                f.write(mod_str(testcase.name) + ',' + mod_str(testcase.script_name) + ',')
                
                temp = ""
                for var in testcase.variables.all():
                    temp += '('+var.name+','+var.value+'),'
                temp = temp.strip(',')
                f.write(mod_str(temp) + ',')
                
                temp = ""
                for misc in testcase.dependencies.all():
                    temp += misc.name + ','
                temp=temp.strip(',')
                f.write(mod_str(temp) + ',')
                
                temp = ""
                for misc in testcase.script_misc_files.all():
                    temp += misc.path + ','
                temp=temp.strip(',')
                f.write(mod_str(temp) + ',')
                
                temp = ""
                for module in testcase.testmodules.all():
                    temp += module.name + ','
                temp=temp.strip(',')
                f.write(mod_str(temp) + ',')
                
                temp = ""
                for tag in testcase.tags.all():
                    temp += tag.name + ','
                temp=temp.strip(',')
                f.write(mod_str(temp) + ',')
                
                f.write(mod_str(testcase.author.name) + ',' + mod_str(testcase.get_state_display()) + ',' + mod_str(testcase.procedure) + ',' + mod_str(str(testcase.p4_script_version)) + ',' )
                f.write(mod_str(testcase.script_type.interpreter) + ',' + mod_str(str(testcase.duration_seconds)) + ',' + mod_str(str(testcase.defects)) + ',' + mod_str(str(testcase.timeout_seconds)) + ',' + mod_str(str(testcase.mrq_number)) + ',')
                
                temp = ""
                for testset in testcase.testsets.all():
                    temp += testset.name + ','
                temp=temp.strip(',')
                f.write(mod_str(temp) + ',')
                
                f.write('1,' + testcase.get_testtype_display() + ',' + testcase.min_stc_version + ',' + testcase.exec_harness.exec_harness + ',' + testcase.mst.name + ',' + str(testcase.priority) + '\r')
        except:
            raise Http404('Error occured when exporting csv file')
        
    return HttpResponseRedirect('/static/download/temp/csv_export_'+time_str+'.csv')
tc_to_csv.short_description = "Export CSV file from selected test cases"

def setCell(ws, col_idx, row_idx, value):
    from openpyxl.cell import get_column_letter
    col = get_column_letter(col_idx)
    ws.cell('%s%s'%(col, row_idx)).value = value

def tc_to_xlsx(modeladmin, request, queryset):
    from openpyxl import Workbook
    
    time_str = timezone.now().strftime('%Y-%m-%d_%H%M%S%f')
    wb = Workbook()
    dest_filename = '/home/SmartTest/WebINF/stsite/static/download/temp/excel_export_'+time_str+'.xlsx'
    ws = wb.worksheets[0]
    ws.title = "smarttest test cases"
    
    field_list = ['TestName','TestScriptName','Variables', 'Dependencies','ScriptMiscFiles','TestModule','Tags','Author','State','Procedure','P4ScriptVersion','ScriptType','Duration', 'Defects', 'TimeOut','MRQ','TestSet','TimeStamp','TestType','MinSTCVersion','ExecHarness','MST','Priority','Objective','Test Plan Section','Environment','Expected Results','Initial Configuration']
    
    for col_idx in xrange(1, len(field_list) + 1):
        setCell(ws, col_idx, 1, field_list[col_idx-1])
        
    try:
        row_idx = 2
        for testcase in queryset.select_related():
            col_idx = 1
            
            setCell(ws, col_idx, row_idx, testcase.name)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.script_name)
            col_idx += 1
            
            temp = ""
            for var in testcase.variables.all():
                temp += '('+var.name+','+var.value+'),'
            temp = temp.strip(',')
            setCell(ws, col_idx, row_idx, temp)
            col_idx += 1
            
            temp = ""
            for misc in testcase.dependencies.all():
                temp += misc.name + ','
            temp=temp.strip(',')
            setCell(ws, col_idx, row_idx, temp)
            col_idx += 1
                     
            temp = ""
            for misc in testcase.script_misc_files.all():
                temp += misc.path + ','
            temp=temp.strip(',')
            setCell(ws, col_idx, row_idx, temp)
            col_idx += 1
            
            temp = ""
            for module in testcase.testmodules.all():
                temp += module.name + ','
            temp=temp.strip(',')
            setCell(ws, col_idx, row_idx, temp)
            col_idx += 1
            
            temp = ""
            for tag in testcase.tags.all():
                temp += tag.name + ','
            temp=temp.strip(',')
            setCell(ws, col_idx, row_idx, temp)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.author.name)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.get_state_display())
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.procedure)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, str(testcase.p4_script_version))
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.script_type.interpreter)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, str(testcase.duration_seconds))
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, str(testcase.defects))
            col_idx += 1

            setCell(ws, col_idx, row_idx, str(testcase.timeout_seconds))
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, str(testcase.mrq_number))
            col_idx += 1
            
            temp = ""
            for testset in testcase.testsets.all():
                temp += testset.name + ','
            temp=temp.strip(',')
            setCell(ws, col_idx, row_idx, temp)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, '1')
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.get_testtype_display())
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.min_stc_version)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.exec_harness.exec_harness)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.mst.name)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, str(testcase.priority))
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.objective)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.test_plan_section)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.environment)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.expected_results)
            col_idx += 1
            
            setCell(ws, col_idx, row_idx, testcase.initial_configuration)
            col_idx += 1
            
            row_idx += 1
            
        wb.save(filename = dest_filename)
            
    except:
        print traceback.format_exc()
        raise Http404('Error occured when exporting spreadsheet')
    
    return HttpResponseRedirect('/static/download/temp/excel_export_'+time_str+'.xlsx')
tc_to_xlsx.short_description = "Export Excel spreadsheet from selected test cases"

def tc_to_text(modeladmin, request, queryset):
    response = HttpResponse()
    for testcase in queryset:
        response.write("<p>" + testcase.name + "</p>")
    return response
tc_to_text.short_description = "Get text view of selected test cases"


def duplicate_tc(modeladmin, request, queryset):
    for testcase in queryset:
        tc_tags = testcase.tags
        tc_vars = testcase.variables
        tc_miscs = testcase.misc_file_location
        tc_testsets = testcase.testsets
        tc_modules = testcase.testmodules
        tc_dependencies = testcase.dependencies
        tc_script_misc_files = testcase.script_misc_files
        new_name = testcase.name + '_Copy'
        finish = False
        temps = TestCase.objects.filter(name=new_name)
        if temps is None or len(temps) == 0:
            testcase.name = new_name
            finish = True
        count = 1
        while not finish:
            new_name = testcase.name + '_Copy_' + str(count)
            temps = TestCase.objects.filter(name=new_name)
            if temps is None or len(temps) == 0:
                testcase.name = new_name
                finish = True
            count +=1
        testcase.pk = None
        testcase.save()
        testcase.tags = tc_tags.all()
        testcase.variables = tc_vars.all()
        testcase.misc_file_location = tc_miscs.all()
        testcase.testsets = tc_testsets.all()
        testcase.testmodules = tc_modules.all()
        testcase.dependencies = tc_dependencies.all()
        testcase.script_misc_files = tc_script_misc_files.all()
        testcase.save()
duplicate_tc.short_description = "Duplicate selected testcases"

def delete_selected_rev(self, request, queryset):
    for testcase in queryset:
        testcase.save()
    queryset.all().delete()
delete_selected_rev.short_description = "Delete selected test cases"

def mass_edit_tc(self, request, queryset):
    tc_ids = ''
    for testcase in queryset:
        tc_ids += '%d,' % (testcase.id)
    return HttpResponseRedirect('/stadmin/stapp/testcase/select_fields/%s' % (tc_ids.strip(',')))
mass_edit_tc.short_description = "Edit selected test cases"

def revert_selected(self, request, queryset):
    for rev in queryset:
        vers = reversion.models.Version.objects.filter(revision_id=rev.id)
        for ver in vers:
            if ver.type == 2:
                messages.error(request, "Cannot revert to a delete action")
                return
    for rev in queryset:
        rev.revert()
revert_selected.short_description = "Revert selected revisions"

class authorInline(admin.StackedInline):
    model = TestCase

class BaseVariableFormSet(BaseFormSet):
    def clean(self):
        if any(self.errors):
            return
        for i in range(0, self.total_form_count()):
            form = self.forms[i]
            val = form.cleaned_data['value']
            if val == '' or val is None:
                raise forms.ValidationError("Variables are empty")

class VariableForm(forms.Form):
    value = forms.CharField(max_length=200, label='Variable')

    def __init__(self, *args, **kwargs):
        global CURR_VAR_ID
        print 'current var id'
        print CURR_VAR_ID
        super(VariableForm, self).__init__(*args, **kwargs)
        self.empty_permitted = False
        if TC_ID != '':
            tc = TestCase.objects.get(pk=TC_ID)
            if tc.exec_harness.exec_harness == 'thot' and tc.script_type.interpreter == 'tcl':
                try:
                    script = TestScript.objects.get(name=tc.script_name.strip())
                    variable = script.variables.split(',')[CURR_VAR_ID]
                    variable_value = ''
                    variable_value = tc.variables.all().get(name=variable).value
                except ObjectDoesNotExist:
                    print 'Variable ' + variable + ' does not exist'
                except:
                    print 'unknown variable error for' + variable
            else:
                count = 0
                variable = ''
                variable_value = ''
                for v in tc.variables.all():
                    if count == CURR_VAR_ID:
                        variable = v.name
                        variable_value = v.value
                        break
                    count += 1
            self.fields['value'].label = variable
            self.fields['value'].initial = variable_value
            CURR_VAR_ID += 1

    def clean_value(self):
        var_value = self.cleaned_data['value']
        if var_value is None or var_value == '':
            raise forms.ValidationError('Varibale value is empty')
        return var_value
        
class variablesInline(admin.StackedInline):
    model = TestCase.variables.through
    extra = 2

class TagFilter(SimpleListFilter):
    title = _('tags')
    parameter_name = 'tag'

    def lookups(self, request, model_admin):
        if request.user.is_superuser:
            tags = Tag.objects.all().order_by('name')
        else:
            tags = TagManager.objects.get(user=request.user.id).tags.all().order_by('name')
        return [(tag.name, tag) for tag in tags]

    def queryset(self, request, queryset):
        if self.value() is not None:
            if request.user.is_superuser:
                tag = Tag.objects.get(name=self.value())
            else:
                tag = TagManager.objects.get(user=request.user.id).tags.get(name=self.value())
            return queryset.filter(tags__in=[tag]).distinct()
        else:
            return queryset.all()

class ErrorFilter(SimpleListFilter):
    title = _('has validation error')
    parameter_name = 'hasvalidationerror'
    def lookups(self, request, model_admin):
        return [('yes', 'yes'), ('no','no')]

    def queryset(self, request, queryset):
        if self.value() == 'yes':
            tc_list = []
            for tc in queryset.all():
                if tc.colored_name().find('<span title=') > -1:
                    tc_list.append(tc.id)
            return queryset.filter(id__in=tc_list)
        elif self.value() == 'no':
            tc_list = []
            for tc in queryset.all():
                if tc.colored_name().find('<span title=') < 0:
                    tc_list.append(tc.id)
            return queryset.filter(id__in=tc_list)
        else:
            return queryset.all()

##class userFilter(SimpleListFilter):
##    title = _('modified by')
##    parameter_name = 'user'
##
##    def lookups(self, request, model_admin):
##        return ['All', 'Me',]
##
##    def queryset(self, request, queryset):
##        if self.value() is not None:
##            if self.value == 'All':
##                return queryset.all()
##            elif self.value == 'Me':
##                les = LogEntry.objects.filter(user=request.user.id)
##                for le in les:
##                    if le.content_type ==
##        return queryset.all()
    
special_fields = ['min_stc_version','procedure','expected_results','initial_configuration','environment','objective']

class MyTestCaseAdmin(forms.ModelForm):
    class Meta:
        model = TestCase

    author_name = forms.CharField(required=False)
    tc_obj = None

    #script_variables = forms.CharField(help_text='Example: (variable1,value1),(variable2,value2)', widget=forms.Textarea, required=False)

    tuple_list = ['NoVar']

    def str_to_tuple_list(self, csv_str, lower_case=None):
        ret = 0
        oc = cc = 0
        self.tuple_list = []
        found = False
        expect = False
        while True:
            oc = csv_str.find('(', cc)
            if oc == -1:
                break
            else:
                expect = True
            
            cc = csv_str.find(')', oc)
            if cc == -1:
                break
            else:
                expect = False
                found = True

            tuple_text = csv_str[oc+1:cc]

            if ',' in tuple_text:
                n,v = tuple_text.split(',', 1)
            elif ' ' in tuple_text:
                n,v = tuple_text.split(' ', 1)
            else:
                print 'ERROR: cannot parse variable:', tuple_text
                #continue
                ret = -1
                break

            if lower_case is None:
                self.tuple_list.append((n.strip(), v.strip()))
            elif lower_case is True:
                self.tuple_list.append((n.strip().lower(), v.strip().lower()))
            else:
                self.tuple_list.append((n.strip().upper(), v.strip().upper()))
        if ret == 0:
            if ( not found ) and csv_str != '':
                ret = -1
            if expect:
                ret = -1
        return ret

    def __init__(self, *args, **kwargs):
        global SELECTED_SCRIPT
        self.request = kwargs.pop('request', None)
        super(MyTestCaseAdmin, self).__init__(*args, **kwargs)

        if 'instance' in kwargs:
            self.initial['author_name'] = kwargs['instance'].author
            self.vars = kwargs['instance'].variables
            self.tc_obj = kwargs['instance']
            #del self.fields['script_variables']

##    def clean_script_name(self):
##        script_name = self.cleaned_data['script_name']
##        try:
##            script = TestScript.objects.get(name=script_name)
##        except ObjectDoesNotExist:
##            print 'The script is not in Perforce'
##            raise forms.ValidationError(mark_safe('The script is not in Perforce. Please submit the script to Perforce and use \"<a href=\"/stapp/syncp4/\">Sync Perforce</a>\" tool'))
##        return script_name

    def clean_author_name(self):
        if require_author:
            author_name = self.cleaned_data['author_name']
            if author_name is None or author_name.strip() == '':
                print 'author_name is required'
                raise forms.ValidationError('This field is required')
            author = TestCaseAuthor.objects.get_or_create(
                name=author_name.strip().lower(),
                defaults={'last_name':'', 'first_name':'','email':''})[0]
            return author
        else:
            print 'author not required'
            return ''

##    def clean_script_variables(self):
##        variable_tex = self.cleaned_data['script_variables']
##        if variable_tex == '' or variable_tex is None:
##            return variable_tex
##        if self.str_to_tuple_list(variable_tex) != 0:
##            print 'Variable format error'
##            raise forms.ValidationError('Variable format error')
##        return variable_tex
        
    def clean_name(self):
        test_name = self.cleaned_data['name']
        if 'name' not in self.changed_data:
            return test_name.strip().lower()
        
        #deny special characters
        match = re.match(r"^[a-zA-Z0-9\_\-\.]+$",test_name)
        if not match:
            raise forms.ValidationError('Test name can only contain a-z, A-z, 0-9, "_", "-", and "."')
        
        duplicate = 0
        if len(test_name) <= 200:
            if self.tc_obj is None:
                duplicate = TestCase.objects.filter(name__iexact=test_name.strip().lower()).count()
            tname=test_name.strip().lower()
        else:
            print 'Test name is too long. 200 characters is the limit'
            raise forms.ValidationError('Test name is too long. 200 characters is the limit')

        if duplicate > 0:
            print 'Test case name already exists in the database'
            raise forms.ValidationError('Test case name already exists in the database')

        return tname
    
    def clean_state(self):
        st = self.cleaned_data['state']
        return st
    
    def clean(self):
		global VAR_EMPTY
		cleaned_data = super(MyTestCaseAdmin, self).clean()
		duration_seconds = cleaned_data.get("duration_seconds")
		timeout_seconds = cleaned_data.get("timeout_seconds")
		new_state = cleaned_data.get("state")
		
		if duration_seconds is not None and timeout_seconds is not None:
			try:
				int(duration_seconds)
			except:
				print "Duration is not number"
				raise forms.ValidationError("Duration is not number")
			try:
				int(timeout_seconds)
			except:
				print "Timeout is not number"
				raise forms.ValidationError("Timeout is not number")
			if int(duration_seconds) > int(timeout_seconds):
				print "Error: Timeout is shorter than duration!"
				raise forms.ValidationError("Error: Timeout is shorter than duration!")
##        if VAR_EMPTY:
##            VAR_EMPTY = False
##            print "Error: Empty Variables"
##            raise forms.ValidationError("Error: Empty Variables")
        
		
		#get variable:locked, adminlocked from stapp_db_authenticate		
		if new_state==0:
			try:
				adminlocked = DBFlag.objects.get(name='adminlocked').value
			except:
				print "Error when get adminlocked, set to default value: false"
				adminlocked = False
				
			try:
				locked = DBFlag.objects.get(name='locked').value
			except:
				print "Error when get locked, set to default value: false"
				locked = False
			
			more_than_special = set(self.changed_data)-set(special_fields)
			if not self.request.user.is_superuser and locked and more_than_special:
				print "SmartTest DB Interface has been locked for user, can't change testcase's state to active"
				raise forms.ValidationError("SmartTest DB Interface has been locked for user, can't change testcase's state to active")
			elif self.request.user.is_superuser and adminlocked and more_than_special:
				print "SmartTest DB Interface has been locked for superuser, can't change testcase's state to active"
				raise forms.ValidationError("SmartTest DB Interface has been locked for superuser, can't change testcase's state to active")
			
			
		if self.request.method == 'POST':
			exist_vars = []
			for postItem in self.request.POST.items():
				if postItem[0].startswith('var_value_'):
					var_name = postItem[0].replace('var_value_','',1)
					exist_vars.append(var_name.strip())
##                    if postItem[1].strip() == '' and self.request.POST.get('var_del_'+var_name) is None:
##                        raise forms.ValidationError("Variable value cannot be blank")
			nvar_names = self.request.POST.getlist('nvar_name')
			nvar_values = self.request.POST.getlist('nvar_value')
			for i in range(0,len(nvar_names)):
				if nvar_names[i].strip() == '' and nvar_values[i].strip() != '':
					raise forms.ValidationError("Variable name cannot be blank")
				if nvar_names[i].strip() != '' and nvar_values[i].strip() != '':
					exist_vars.append(nvar_names[i].strip())
		return cleaned_data


    def save(self, commit=True):
        model = super(MyTestCaseAdmin, self).save(commit=False)
	if model.defects != "" and model.state == 0:
		model.state = 5
        
        if require_author:
            model.author = self.cleaned_data['author_name']
            
        # get variables
        var_to_add = []
        if self.request.method == 'POST':
            for postItem in self.request.POST.items():
                if postItem[0].startswith('var_value_'):
                    var_name = postItem[0].replace('var_value_','',1)
                    if self.request.POST.get('var_del_'+var_name) is None:
                        try:
                            var_to_add.append(Variable.objects.get_or_create(name=var_name.strip(),value=postItem[1].strip())[0])
                        except:
                            print 'Error adding variable ' + var_name
            nvar_names = self.request.POST.getlist('nvar_name')
            nvar_values = self.request.POST.getlist('nvar_value')
            for i in range(0,len(nvar_names)):
                try:
                    if nvar_names[i].strip():
                        var_to_add.append(Variable.objects.get_or_create(name=nvar_names[i].strip(),value=nvar_values[i].strip())[0])
                except:
                    print 'Error adding variable ' + nvar_names[i]
        
        # get script misc files
        misc_to_add = []
        if self.request.method == 'POST':
            miscs = self.request.POST.getlist('scriptmiscfiles_value')
            if miscs is None:
                miscs = []
            for misc in miscs:
                if misc.strip() == '':
                    continue
                misc_obj = None
                misc_qstemp = ScriptMiscFile.objects.filter(path=misc)
                if misc_qstemp.count() == 0:
                    misc_obj = ScriptMiscFile(path = misc)
                    misc_obj.save()
                else:
                    misc_obj = misc_qstemp[0]
                misc_to_add.append(misc_obj)
        
        # check if it is add
        temp = self.request.get_full_path().split('/')
        tc_id = temp[len(temp)-2]
        isAdd = (tc_id == 'add') or ('recover' in self.request.get_full_path())
        
        # check if it is mass edit
        isMassEdit = ('mass_edit' in temp)
        
        # get old state
        old_obj = None
        old_state = 2
        if not isAdd:
            old_obj = TestCase.objects.get(id=model.id)
            old_state = old_obj.state
            isVariableChanged=False
            if not isMassEdit and set(var_to_add) != set(old_obj.variables.all()):
                isVariableChanged=True
                self.changed_data.append('variable')
        
            isMiscChanged=False
            if not isMassEdit and set(filter(None, miscs)) != set(old_obj.script_misc_files.all().values_list('path',flat=True)):
                isMiscChanged=True
                self.changed_data.append('script_misc_file')
        
        # check if data is changed
        isChanged = True
        if ( not isAdd ):
            isChanged = False
            isChanged = isChanged or (len(set(self.changed_data)-set(special_fields))>0)
            isChanged = isChanged or isVariableChanged or isMiscChanged
            #isChanged = isChanged or (model.author != old_obj.author)
        
        if self.request.user.is_superuser:
            try:
                if model.previous_version is not None:
                    model.previous_version.state = 2
                    model.previous_version.save()
            except:
                pass
        else:
            stateMan = StateManager(model, old_state, isChanged, isAdd)
            model = stateMan.changeState()
        
        if not isMassEdit and self.request.POST.get('script_name') is not None:
	    #raise Http404(old_obj.defects)
	    #raise Http404(model.state)
            model.save()
            model.variables = var_to_add
            model.script_misc_files = misc_to_add
        
#        
#        # create new version if there is a change and previous state is active
#        isNewRow = False
#        if not isAdd and isChanged and (old_obj.state == 0 and model.state!=1):
#            model.pk = None
#            model.save()
#            isNewRow = True
        
#        # load m2m fields from old version
#        if ( not isAdd ):
#            for attr in dir(old_obj):
#                if attr == 'objects':
#                    continue
#                attr_obj = old_obj.__getattribute__(attr)
#                if (str(type(attr_obj)) == "<class 'django.db.models.fields.related.ManyRelatedManager'>"):
#                    model.__setattr__(attr,attr_obj.all())
#            model.save()
        
#        # change state based on change
#        if isAdd:
#            model.state = 3
#        elif isChanged and isNewRow and model.state != 2:
#            model.state = 3
#        else:
#            pass
            
#        # save variables
#        if not isMassEdit and self.request.POST.get('script_name') is not None:
#            model.save()
#            model.variables = var_to_add
#
#        # depreacate old version if promote or deprecate
#        if not isAdd:
#            if (not isNewRow) and model.state==0:
#                try:
#                    pre_ver = TestCase.objects.get(id=model.previous_version.id)
#                    pre_ver.state = 2
#                    pre_ver.save()
#                except:
#                    pass
#            if isNewRow and old_obj is not None and model.state==2:
#                old_obj.state = 2
#                old_obj.save()
                
#        # link versions
#        if not isAdd and old_obj is not None and isNewRow:
#            model.previous_version = old_obj
            
        model.save()
 
        return model

class FieldSelForm(forms.Form):
    script_name = forms.BooleanField(required=False, label="Script name")
    #script_variables = forms.BooleanField(required=False, label="Script variables")
    #misc_file_location = forms.BooleanField(required=False, label="Miscellaneous file locations")
    testmodules = forms.BooleanField(required=False, label="Test modules")
    tags = forms.BooleanField(required=False, label="Tags")
    author_name = forms.BooleanField(required=False, label="Author")
    state = forms.BooleanField(required=False, label="State")
    procedure = forms.BooleanField(required=False, label="Procedure")
    p4_script_version = forms.BooleanField(required=False, label="P4 script version")
    script_type = forms.BooleanField(required=False, label="script_type")
    duration_seconds = forms.BooleanField(required=False, label="Duration")
    defects = forms.BooleanField(required=False, label="Defects")
    timeout_seconds = forms.BooleanField(required=False, label="Timeout")
    mrq_number = forms.BooleanField(required=False, label="MRQ number")
    testsets = forms.BooleanField(required=False, label="Test sets")
    testtype = forms.BooleanField(required=False, label="Test type")
    min_stc_version = forms.BooleanField(required=False, label="Min STC version")
    exec_harness = forms.BooleanField(required=False, label="Execution harness")
    mst = forms.BooleanField(required=False, label="MST")
    priority = forms.BooleanField(required=False, label="Priority")
    test_plan_section = forms.BooleanField(required=False, label="Test Plan Section")
    environment = forms.BooleanField(required=False, label="Environment")
    objective = forms.BooleanField(required=False, label="Objective")
    expected_results = forms.BooleanField(required=False, label="Expected Results")
    initial_configuration = forms.BooleanField(required=False, label="Initial Configuration")
    dependencies = forms.BooleanField(required=False, label="Dependencies")
    #script_misc_files = forms.BooleanField(required=False, label="Script_Misc_Files")

class TestCaseAdmin(reversion.VersionAdmin):

    def get_form(self, request, obj=None, **kwargs):
        AdminForm = super(TestCaseAdmin, self).get_form(request, obj, **kwargs)
        class ModelFormMetaClass(AdminForm):
            def __new__(cls, *args, **kwargs):
                kwargs['request'] = request
                return AdminForm(*args, **kwargs)
        return ModelFormMetaClass

    def lookup_allowed(self, lookup, value):
        return True

    @csrf_protect_m
    @transaction.commit_on_success
    def mass_change_view(self, request, object_id_list, form_url='', extra_context={'isMass':True}):
        global require_author
        require_author = request.session.get('require_author')
        "The 'change' admin view for this model."
        model = self.model
        opts = model._meta

        object_ids = object_id_list.split(',')

        for tc_id in object_ids:
            try:
                obj = self.queryset(request).get(pk=unquote(tc_id))
            except model.DoesNotExist:
                obj = None
            if not self.has_change_permission(request, obj):
                raise PermissionDenied
            if obj is None:
                raise Http404(_('%(name)s object with primary key %(key)r does not exist.') % {'name': force_unicode(opts.verbose_name), 'key': escape(tc_id)})
            

##        if request.method == 'POST' and "_saveasnew" in request.POST:
##            return self.add_view(request, form_url=reverse('admin:%s_%s_add' %
##                                    (opts.app_label, opts.module_name),
##                                    current_app=self.admin_site.name))

        obj = self.queryset(request).get(pk=unquote(object_ids[0]))
        
        ModelForm = self.get_form(request, obj)
        formsets = []
        inline_instances = self.get_inline_instances(request)
        if request.method == 'POST':
            with transaction.commit_manually():
                try:
                    objects = self.queryset(request).filter(pk__in = object_ids)
                    allvalid = True
                    for obj in objects:
                        form = ModelForm(request.POST, request.FILES, instance=obj)
                        #del form.fields['name']
                        exclude = request.session.get('to_exclude')
                        for fld in request.session.get('to_exclude'):
                            print fld
                            del form.fields[fld]
                        if form.is_valid():
                            form_validated = True
                            new_object = self.save_form(request, form, change=True)
                            #new_object.name = obj.name
                        else:
                            form_validated = False
                            new_object = obj
                        prefixes = {}
                        for FormSet, inline in zip(self.get_formsets(request, new_object), inline_instances):
                            prefix = FormSet.get_default_prefix()
                            prefixes[prefix] = prefixes.get(prefix, 0) + 1
                            if prefixes[prefix] != 1 or not prefix:
                                prefix = "%s-%s" % (prefix, prefixes[prefix])
                            formset = FormSet(request.POST, request.FILES,
                                              instance=new_object, prefix=prefix,
                                              queryset=inline.queryset(request))

                            formsets.append(formset)

                        if all_valid(formsets) and form_validated:
                            self.save_model(request, new_object, form, True)
                            self.save_related(request, form, formsets, True)
                            change_message = self.construct_change_message(request, form, formsets)
                            self.log_change(request, new_object, change_message)
                            #return self.response_change(request, new_object)
                        else:
                            allvalid = False
                    #transaction.commit()
                    if allvalid:
                        request.session['require_author'] = True
                        require_author = True
                        transaction.commit()
                        return self.response_change(request, new_object)
                    else:
                        #pass
                        transaction.rollback()
                except:
                    request.session['require_author'] = True
                    print traceback.format_exc()
                    transaction.rollback()

        else:
            initial = dict(request.GET.items())
            for k in initial:
                try:
                    f = opts.get_field(k)
                except models.FieldDoesNotExist:
                    continue
                if isinstance(f, models.ManyToManyField):
                    initial[k] = initial[k].split(",")
            form = ModelForm(initial=initial)
            prefixes = {}
            for FormSet, inline in zip(self.get_formsets(request, obj), inline_instances):
                prefix = FormSet.get_default_prefix()
                prefixes[prefix] = prefixes.get(prefix, 0) + 1
                if prefixes[prefix] != 1 or not prefix:
                    prefix = "%s-%s" % (prefix, prefixes[prefix])
                formset = FormSet(instance=obj, prefix=prefix,
                                  queryset=inline.queryset(request))
                formsets.append(formset)

        adminForm = helpers.AdminForm(form, self.get_fieldsets(request, obj),
            self.get_prepopulated_fields(request, obj),
            self.get_readonly_fields(request, obj),
            model_admin=self)
        media = self.media + adminForm.media

        inline_admin_formsets = []
        for inline, formset in zip(inline_instances, formsets):
            fieldsets = list(inline.get_fieldsets(request, obj))
            readonly = list(inline.get_readonly_fields(request, obj))
            prepopulated = dict(inline.get_prepopulated_fields(request, obj))
            inline_admin_formset = helpers.InlineAdminFormSet(inline, formset,
                fieldsets, prepopulated, readonly, model_admin=self)
            inline_admin_formsets.append(inline_admin_formset)
            media = media + inline_admin_formset.media

        context = {
            'title': _('Change %s') % force_unicode(opts.verbose_name),
            'adminform': adminForm,
            'object_id': object_ids,
            'original': obj,
            'is_popup': "_popup" in request.REQUEST,
            'media': media,
            'inline_admin_formsets': inline_admin_formsets,
            'errors': helpers.AdminErrorList(form, formsets),
            'app_label': opts.app_label,
        }
        context.update(extra_context or {})
        return self.render_change_form(request, context, change=True, obj=obj, form_url=form_url)

    def mass_edit_view(self, request, object_ids):
        #return self.mass_change_view(request, object_ids)
        result = self.mass_change_view(request, object_ids)
        
        ref = request.META.get('HTTP_REFERER', '')
        if ref.find('?') != -1:
            request.session['filtered'] =  ref
        
        if request.POST.has_key('_save'):
            try:
                if request.session['filtered'] is not None:
                    result['Location'] = request.session['filtered']
                    request.session['filtered'] = None
            except:
                pass
                
        return result

    def select_fields_view(self, request, object_ids):
        global to_edit, to_exclude, require_author
        to_edit = []
        to_exclude = ['name',]
        request.session['to_edit'] = []
        request.session['to_exclude'] = ['name',]
        request.session['require_author'] = True
        if request.method == 'POST':
            form = FieldSelForm(request.POST)
            to_edit = []
            to_exclude = ['name',]
            if form.is_valid():
                
                if form.cleaned_data['script_name']:
                    to_edit.append('script_name')
                else:
                    to_exclude.append('script_name')
                    
                #if form.cleaned_data['script_variables']:
                    #to_edit.append('script_variables')
                #else:
                    #to_exclude.append('script_variables')

#                if form.cleaned_data['misc_file_location']:
#                    to_edit.append('misc_file_location')
#                else:
#                    to_exclude.append('misc_file_location')

                if form.cleaned_data['testmodules']:
                    to_edit.append('testmodules')
                else:
                    to_exclude.append('testmodules')

                if form.cleaned_data['tags']:
                    to_edit.append('tags')
                else:
                    to_exclude.append('tags')

                if form.cleaned_data['author_name']:
                    to_edit.append('author_name')
                else:
                    to_exclude.append('author_name')
                    require_author = False

                if form.cleaned_data['state']:
                    to_edit.append('state')
                else:
                    to_exclude.append('state')

                if form.cleaned_data['procedure']:
                    to_edit.append('procedure')
                else:
                    to_exclude.append('procedure')

                if form.cleaned_data['p4_script_version']:
                    to_edit.append('p4_script_version')
                else:
                    to_exclude.append('p4_script_version')

                if form.cleaned_data['script_type']:
                    to_edit.append('script_type')
                else:
                    to_exclude.append('script_type')

                if form.cleaned_data['duration_seconds']:
                    to_edit.append('duration_seconds')
                else:
                    to_exclude.append('duration_seconds')

                if form.cleaned_data['defects']:
                    to_edit.append('defects')
                else:
                    to_exclude.append('defects')

                if form.cleaned_data['timeout_seconds']:
                    to_edit.append('timeout_seconds')
                else:
                    to_exclude.append('timeout_seconds')

                if form.cleaned_data['mrq_number']:
                    to_edit.append('mrq_number')
                else:
                    to_exclude.append('mrq_number')

                if form.cleaned_data['testsets']:
                    to_edit.append('testsets')
                else:
                    to_exclude.append('testsets')

                if form.cleaned_data['testtype']:
                    to_edit.append('testtype')
                else:
                    to_exclude.append('testtype')

                if form.cleaned_data['min_stc_version']:
                    to_edit.append('min_stc_version')
                else:
                    to_exclude.append('min_stc_version')

                if form.cleaned_data['exec_harness']:
                    to_edit.append('exec_harness')
                else:
                    to_exclude.append('exec_harness')
                
                if form.cleaned_data['mst']:
                    to_edit.append('mst')
                else:
                    to_exclude.append('mst')

                if form.cleaned_data['priority']:
                    to_edit.append('priority')
                else:
                    to_exclude.append('priority')
                    
                if form.cleaned_data['objective']:
                    to_edit.append('objective')
                else:
                    to_exclude.append('objective')
                    
                if form.cleaned_data['test_plan_section']:
                    to_edit.append('test_plan_section')
                else:
                    to_exclude.append('test_plan_section')
                
                if form.cleaned_data['environment']:
                    to_edit.append('environment')
                else:
                    to_exclude.append('environment')
                    
                if form.cleaned_data['expected_results']:
                    to_edit.append('expected_results')
                else:
                    to_exclude.append('expected_results')
                
                if form.cleaned_data['initial_configuration']:
                    to_edit.append('initial_configuration')
                else:
                    to_exclude.append('initial_configuration')
                
                if form.cleaned_data['dependencies']:
                    to_edit.append('dependencies')
                else:
                    to_exclude.append('dependencies')
                    
#                if form.cleaned_data['script_misc_files']:
#                    to_edit.append('script_misc_files')
#                else:
#                    to_exclude.append('script_misc_files')

##                self.fieldsets = (
##                    (None, {'fields': to_edit}),
##                    )
                #self.exclude = to_exclude

                print to_edit
                request.session['to_edit'] = to_edit
                request.session['to_exclude'] = to_exclude
                request.session['require_author'] = require_author

                return HttpResponseRedirect('/stadmin/stapp/testcase/mass_edit/%s' % (object_ids))
                    
        else:
            form = FieldSelForm()
            
        ref = request.META.get('HTTP_REFERER', '')
        if ref.find('?') != -1:
            # We've got a query string, set the session value
            request.session['filtered'] =  ref
        
        return render_to_response('admin/select_fields.html',
            {'form': form},
            context_instance=RequestContext(request)
        )

    def adv_filter(self, request, query):
        request.session['q_adv'] = query
        return HttpResponseRedirect('/stadmin/stapp/testcase/adv_fil/')

    def adv_filter_totc(self, request, tc_id):
        str_url = '/stadmin/stapp/testcase/' + tc_id + '/'
        return HttpResponseRedirect(str_url)
    
    def write_q_str(self, q_str, field_name, request):
        l_kw = request.POST.getlist(field_name + '_keyword')
        l_con = request.POST.getlist(field_name + '_con')
        field_log = request.POST[field_name +'_connect']
        q_field = ''
        for i in range(0,len(l_con)):
            if l_con[i] == 'default' or l_kw[i] is None:
                continue
            q_field += l_con[i] + l_kw[i] + ';'
        q_field = q_field.strip(';')
        if q_field != '':
            q_str += field_name + '_' + field_log + '=' + q_field + ':'
        return q_str
    
    def adv_filter_sel(self, request):
        if request.method == 'POST':
            q_str = ''
            
            # get name filter
            l_name_kw = request.POST.getlist('name_keyword')
            l_name_con = request.POST.getlist('name_con')
            name_log = request.POST['name_connect']
            q_name = ''
            for i in range(0,len(l_name_con)):
                if l_name_con[i] == 'default' or l_name_kw[i] is None or l_name_kw[i] == '':
                    continue
                q_name += l_name_con[i] + l_name_kw[i] + ';'
            q_name = q_name.strip(';')
            if q_name != '':
                q_str += 'nm_' + name_log + '=' + q_name + ':'

			# get script filter
            l_script_kw = request.POST.getlist('script_keyword')
            l_script_con = request.POST.getlist('script_con')
            script_log = request.POST['script_connect']
            q_script = ''
            for i in range(0,len(l_script_con)):
                if l_script_con[i] == 'default' or l_script_kw[i] is None or l_script_kw[i] == '':
                    continue
                q_script += l_script_con[i] + l_script_kw[i] + ';'
            q_script = q_script.strip(';')
            if q_script != '':
                q_str += 'script_' + script_log + '=' + q_script + ':'

            # get testset filter
            l_ts_kw = request.POST.getlist('ts_keyword')
            l_ts_con = request.POST.getlist('ts_con')
            ts_log = request.POST['testset_connect']
            q_ts = ''
            for i in range(0,len(l_ts_con)):
                if l_ts_con[i] == 'default' or l_ts_kw[i] is None or l_ts_kw[i] == '':
                    continue
                q_ts += l_ts_con[i] + l_ts_kw[i] + ';'
            q_ts = q_ts.strip(';')
            if q_ts != '':
                q_str += 'ts_' + ts_log + '=' + q_ts + ':'

            # get testtype filter
            l_testtype_kw = request.POST.getlist('testtype_keyword')
            l_testtype_con = request.POST.getlist('testtype_con')
            testtype_log = request.POST['testtype_connect']
            q_testtype = ''
            for i in range(0,len(l_testtype_con)):
                if l_testtype_con[i] == 'default' or l_testtype_kw[i] is None or l_testtype_kw[i] == '':
                    continue
                q_testtype += l_testtype_con[i] + l_testtype_kw[i] + ';'
            q_testtype = q_testtype.strip(';')
            if q_testtype != '':
                q_str += 'testtype_' + testtype_log + '=' + q_testtype + ':'

            # get module filter
            l_tm_kw = request.POST.getlist('tm_keyword')
            l_tm_con = request.POST.getlist('tm_con')
            tm_log = request.POST['module_connect']
            q_tm = ''
            for i in range(0,len(l_tm_con)):
                if l_tm_con[i] == 'default' or l_tm_kw[i] is None or l_tm_kw[i] == '':
                    continue
                q_tm += l_tm_con[i] + l_tm_kw[i] + ';'
            q_tm = q_tm.strip(';')
            if q_tm != '':
                q_str += 'tm_' + tm_log + '=' + q_tm + ':'

			# get author filter
            l_author_kw = request.POST.getlist('author_keyword')
            l_author_con = request.POST.getlist('author_con')
            author_log = request.POST['author_connect']
            q_author = ''
            for i in range(0,len(l_author_con)):
                if l_author_con[i] == 'default' or l_author_kw[i] is None or l_author_kw[i] == '':
                    continue
                q_author += l_author_con[i] + l_author_kw[i] + ';'
            q_author = q_author.strip(';')
            if q_author != '':
                q_str += 'author_' + author_log + '=' + q_author + ':'
			
            # get tag filter
            l_tg_kw = request.POST.getlist('tg_keyword')
            l_tg_con = request.POST.getlist('tg_con')
            tg_log = request.POST['tag_connect']
            q_tg = ''
            for i in range(0,len(l_tg_con)):
                if l_tg_con[i] == 'default' or l_tg_kw[i] is None or l_tg_kw[i] == '':
                    continue
                q_tg += l_tg_con[i] + l_tg_kw[i] + ';'
            q_tg = q_tg.strip(';')
            if q_tg != '':
                q_str += 'tg_' + tg_log + '=' + q_tg + ':'
                
            # get mst filter
            l_ms_kw = request.POST.getlist('ms_keyword')
            l_ms_con = request.POST.getlist('ms_con')
            ms_log = request.POST['mst_connect']
            q_ms = ''
            for i in range(0,len(l_ms_con)):
                if l_ms_con[i] == 'default' or l_ms_kw[i] is None or l_ms_kw[i] == '':
                    continue
                q_ms += l_ms_con[i] + l_ms_kw[i] + ';'
            q_ms = q_ms.strip(';')
            if q_ms != '':
                q_str += 'ms_' + ms_log + '=' + q_ms + ':'

            # get duration filter
            l_duration_kw = request.POST.getlist('duration_keyword')
            l_duration_con = request.POST.getlist('duration_con')
            duration_log = request.POST['duration_connect']
            q_duration = ''
            for i in range(0,len(l_duration_con)):
                if l_duration_con[i] == 'default' or l_duration_kw[i] is None or l_duration_kw[i] == '':
                    continue
                q_duration += l_duration_con[i] + l_duration_kw[i] + ';'
            q_duration = q_duration.strip(';')
            if q_duration != '':
                q_str += 'duration_' + duration_log + '=' + q_duration + ':'

            # get defects filter
            l_defects_kw = request.POST.getlist('defects_keyword')
            l_defects_con = request.POST.getlist('defects_con')
            defects_log = request.POST['defects_connect']
            q_defects = ''
            for i in range(0,len(l_defects_con)):
                if l_defects_con[i] == 'default' or l_defects_kw[i] is None or l_defects_kw[i] == '':
                    continue
                q_defects += l_defects_con[i] + l_defects_kw[i] + ';'
            q_defects = q_defects.strip(';')
            if q_defects != '':
                q_str += 'defects_' + defects_log + '=' + q_defects + ':'

            # get timeout filter
            l_timeout_kw = request.POST.getlist('timeout_keyword')
            l_timeout_con = request.POST.getlist('timeout_con')
            timeout_log = request.POST['timeout_connect']
            q_timeout = ''
            for i in range(0,len(l_timeout_con)):
                if l_timeout_con[i] == 'default' or l_timeout_kw[i] is None or l_timeout_kw[i] == '':
                    continue
                q_timeout += l_timeout_con[i] + l_timeout_kw[i] + ';'
            q_timeout = q_timeout.strip(';')
            if q_timeout != '':
                q_str += 'timeout_' + timeout_log + '=' + q_timeout + ':'

            # get priority filter
            l_priority_kw = request.POST.getlist('priority_keyword')
            l_priority_con = request.POST.getlist('priority_con')
            priority_log = request.POST['priority_connect']
            q_priority = ''
            for i in range(0,len(l_priority_con)):
                if l_priority_con[i] == 'default' or l_priority_kw[i] is None or l_priority_kw[i] == '':
                    continue
                q_priority += l_priority_con[i] + l_priority_kw[i] + ';'
            q_priority = q_priority.strip(';')
            if q_priority != '':
                q_str += 'priority_' + priority_log + '=' + q_priority + ':'

            # get mrq filter
            l_mrq_number_kw = request.POST.getlist('mrq_number_keyword')
            l_mrq_number_con = request.POST.getlist('mrq_number_con')
            mrq_number_log = request.POST['mrq_number_connect']
            q_mrq_number = ''
            for i in range(0,len(l_mrq_number_con)):
                if l_mrq_number_con[i] == 'default' or l_mrq_number_kw[i] is None or l_mrq_number_kw[i] == '':
                    continue
                q_mrq_number += l_mrq_number_con[i] + l_mrq_number_kw[i] + ';'
            q_mrq_number = q_mrq_number.strip(';')
            if q_mrq_number != '':
                q_str += 'mrq_number_' + mrq_number_log + '=' + q_mrq_number + ':'
            
            q_str = self.write_q_str(q_str,'objective', request)
            q_str = self.write_q_str(q_str,'initial_configuration', request)
            q_str = self.write_q_str(q_str,'procedure', request)
            q_str = self.write_q_str(q_str,'expected_results', request)
            q_str = self.write_q_str(q_str,'environment', request)
            q_str = self.write_q_str(q_str,'test_plan_section', request)

            q_str = q_str.strip(':')

            if q_str == '':
                q_str = 'all'

            return HttpResponseRedirect('/stadmin/stapp/testcase/adv_fil/' + q_str + '/')
                
        return render_to_response('admin/adv_filter.html', context_instance=RequestContext(request))
    
    def get_urls(self):
        urls = super(TestCaseAdmin,self).get_urls()
        my_urls = patterns('',
            (r'^mass_edit/(?P<object_ids>[0-9,]+)/$', self.mass_edit_view),
            (r'^select_fields/(?P<object_ids>[0-9,]+)/$', self.select_fields_view),
            (r'^adv_fil_sel/$', self.adv_filter_sel),
            (r'^adv_fil/(?P<tc_id>\d+)/$', self.adv_filter_totc),
            (r'^adv_fil/(?P<query>[\S;,%\-\_\.\|\@\ ]+)/$', self.adv_filter),
            (r'^adv_fil/$', self.changelist_view),
        )
        return my_urls + urls
    
    exclude = []

    fieldsets = (
        (None, {'fields': ('name', 'script_name','defects', 'dependencies',
                            'testmodules', 'tags', 'author_name', 'state',
                            'p4_script_version', 'script_type', 'duration_seconds',
                            'timeout_seconds', 'mrq_number', 'testsets', 'testtype',
                            'min_stc_version', 'exec_harness', 'mst', 'priority')}),
        ('TestCase Information',{'classes': ('collapse',),
                        'fields': ('objective','procedure','test_plan_section','environment','expected_results', 'initial_configuration')
                        }),
        )   
    
    def get_fieldsets(self, request, obj=None):
        global to_edit, require_author
        fieldsets = super(TestCaseAdmin, self).get_fieldsets(request, obj)
        temp = request.get_full_path().split('/')
        tc_id = temp[len(temp)-2]
        kw = temp[len(temp)-3]
        if tc_id == 'add':
            require_author = True
            fieldsets = (
                (None, {'fields': ('name', 'script_name',
                                    'testmodules', 'tags', 'author_name', 'state',
                                    'p4_script_version', 'script_type', 'duration_seconds',
                                    'timeout_seconds', 'mrq_number', 'testsets', 'testtype',
                                    'min_stc_version', 'exec_harness', 'mst', 'priority','defects', 'dependencies')}),
                ('TestCase Information',{
                        'fields': ('objective','procedure','test_plan_section','environment','expected_results', 'initial_configuration')
                        }),
                )
        elif kw == 'mass_edit':
            print 'Session Data:'
            print request.session.get('to_edit')
            print request.session.get('to_exclude')
            print request.session.get('require_author')
            fieldsets = (
                (None, {'fields': request.session.get('to_edit')}),
                )
        else:
            require_author = True
            fieldsets = (
                (None, {'fields': ('name', 'script_name',
                                    'testmodules', 'tags', 'author_name', 'state',
                                    'p4_script_version', 'script_type', 'duration_seconds',
                                    'timeout_seconds', 'mrq_number', 'testsets', 'testtype',
                                    'min_stc_version', 'exec_harness', 'mst', 'priority','defects', 'dependencies')}),
                ('TestCase Information',{'classes': ('collapse',),
                        'fields': ('objective','procedure','test_plan_section','environment','expected_results','initial_configuration')
                        }),
                )
        return fieldsets
    
    def q_str_filter(self, q_str, field_name, qs):
        nm_m = re.search(field_name + r"_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
        if nm_m is not None:
            nm_mode = nm_m.group(1)
            nm_list = nm_m.group(2)
            if nm_mode == 'a':
                for nm in nm_list.split(';'):
                    if len(nm) >= 1:
                        match_mode = nm[0]
                        match_content = nm[1:]
                        if match_mode == 'e':
                            kwargs = {'{0}__iexact'.format(field_name):match_content,}
                            qs = qs.filter(**kwargs)
                        elif match_mode == 's':
                            kwargs = {'{0}__istartswith'.format(field_name):match_content,}
                            qs = qs.filter(**kwargs)
                        elif match_mode == 'c':
                            kwargs = {'{0}__icontains'.format(field_name):match_content,}
                            qs = qs.filter(**kwargs)
                        elif match_mode == 'r':
                            kwargs = {'{0}__iexact'.format(field_name):match_content,}
                            qs = qs.exclude(**kwargs)
                        elif match_mode == 'd':
                            kwargs = {'{0}__istartswith'.format(field_name):match_content,}
                            qs = qs.exclude(**kwargs)
                        elif match_mode == 'v':
                            kwargs = {'{0}__icontains'.format(field_name):match_content,}
                            qs = qs.exclude(**kwargs)
            elif nm_mode == 'o':
                qs_temp = qs.none()
                for nm in nm_list.split(';'):
                    if len(nm) >= 1:
                        match_mode = nm[0]
                        match_content = nm[1:]
                        if match_mode == 'e':
                            kwargs = {'{0}__iexact'.format(field_name):match_content,}
                            qs_temp = qs_temp | qs.filter(**kwargs)
                        elif match_mode == 's':
                            kwargs = {'{0}__istartswith'.format(field_name):match_content,}
                            qs_temp = qs_temp | qs.filter(**kwargs)
                        elif match_mode == 'c':
                            kwargs = {'{0}__icontains'.format(field_name):match_content,}
                            qs_temp = qs_temp | qs.filter(**kwargs)
                        elif match_mode == 'r':
                            kwargs = {'{0}__iexact'.format(field_name):match_content,}
                            qs_temp = qs_temp | qs.exclude(**kwargs)
                        elif match_mode == 'd':
                            kwargs = {'{0}__istartswith'.format(field_name):match_content,}
                            qs_temp = qs_temp |  qs.exclude(**kwargs)
                        elif match_mode == 'v':
                            kwargs = {'{0}__icontains'.format(field_name):match_content,}
                            qs_temp = qs_temp |  qs.exclude(**kwargs)
                qs = qs_temp.distinct()
        return qs
        

    def queryset(self, request):
        print request.user.username
        qs = super(TestCaseAdmin, self).queryset(request)
        qs = TestCase.objects.all().exclude(id__in=TestCase.objects.all().exclude(previous_version_id__isnull=True).values_list('previous_version_id',flat=True))
        if request.user.is_superuser:
            #return qs.all()
            print 'super user'
        else:
            tagM = TagManager.objects.get(user=request.user.id)
            qs = qs.filter(tags__in=tagM.tags.all()).distinct()
        
        print 'Query string:'
        print request.session.get('q_adv')
        q_str = request.session.get('q_adv')
        temp = request.get_full_path().split('/')
        if temp[len(temp)-2] == 'adv_fil':
            if q_str == 'all':
                return qs

            nm_m = re.search(r"nm_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if nm_m is not None:
                nm_mode = nm_m.group(1)
                nm_list = nm_m.group(2)
                if nm_mode == 'a':
                    for nm in nm_list.split(';'):
                        if len(nm) > 1:
                            match_mode = nm[0]
                            match_content = nm[1:]
                            print match_content
                            if match_mode == 'e':
                                qs = qs.filter(name__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(name__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(name__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(name__icontains=match_content)
                elif nm_mode == 'o':
                    qs_temp = qs.none()
                    for nm in nm_list.split(';'):
                        if len(nm) > 1:
                            match_mode = nm[0]
                            match_content = nm[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(name__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(name__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(name__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(name__icontains=match_content)
                    qs = qs_temp.distinct()

            script_m = re.search(r"script_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if script_m is not None:
                script_mode = script_m.group(1)
                script_list = script_m.group(2)
                if script_mode == 'a':
                    for script in script_list.split(';'):
                        if len(script) > 1:
                            match_mode = script[0]
                            match_content = script[1:]
                            if match_mode == 'e':
                                qs = qs.filter(script_name__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(script_name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(script_name__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(script_name__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(script_name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(script_name__icontains=match_content)
                elif script_mode == 'o':
                    qs_temp = qs.none()
                    for script in script_list.split(';'):
                        if len(script) > 1:
                            match_mode = script[0]
                            match_content = script[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(script_name__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(script_name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(script_name__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(script_name__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(script_name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(script_name__icontains=match_content)
                    qs = qs_temp.distinct()
            
            ts_m = re.search(r"^ts_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if ts_m is not None:
                ts_mode = ts_m.group(1)
                ts_list = ts_m.group(2)
                if ts_mode == 'a':
                    for ts in ts_list.split(';'):
                        if len(ts) > 1:
                            match_mode = ts[0]
                            match_content = ts[1:]
                            if match_mode == 'e':
                                qs = qs.filter(testsets__name__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(testsets__name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(testsets__name__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(testsets__name__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(testsets__name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(testsets__name__icontains=match_content)
                elif ts_mode == 'o':
                    qs_temp = qs.none()
                    for ts in ts_list.split(';'):
                        if len(ts) > 1:
                            match_mode = ts[0]
                            match_content = ts[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(testsets__name__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(testsets__name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(testsets__name__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(testsets__name__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(testsets__name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(testsets__name__icontains=match_content)
                    qs = qs_temp.distinct()

            testtype_m = re.search(r"testtype_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if testtype_m is not None:
                testtype_mode = testtype_m.group(1)
                testtype_list = testtype_m.group(2)
                if testtype_mode == 'a':
                    for testtype in testtype_list.split(';'):
                        if len(testtype) > 1:
                            match_mode = testtype[0]
                            match_content = testtype[1:]
                            if match_mode == 'e':
                                temp = [k for k,v in dict(TestCase.TESTTYPE_CHOICES).iteritems() if v==match_content]
                                qs = qs.filter(testtype__in=temp)
                            elif match_mode == 's':
                                temp = [k for k,v in dict(TestCase.TESTTYPE_CHOICES).iteritems() if v.startswith(match_content)]
                                qs = qs.filter(testtype__in=temp)
                            elif match_mode == 'c':
                                temp = [k for k,v in dict(TestCase.TESTTYPE_CHOICES).iteritems() if match_content in v]
                                qs = qs.filter(testtype__in=temp)
                            elif match_mode == 'r':
                                temp = [k for k,v in dict(TestCase.TESTTYPE_CHOICES).iteritems() if v==match_content]
                                qs = qs.exclude(testtype__in=temp)
                            elif match_mode == 'd':
                                temp = [k for k,v in dict(TestCase.TESTTYPE_CHOICES).iteritems() if v.startswith(match_content)]
                                qs = qs.exclude(testtype__in=temp)
                            elif match_mode == 'v':
                                temp = [k for k,v in dict(TestCase.TESTTYPE_CHOICES).iteritems() if match_content in v]
                                qs = qs.exclude(testtype__in=temp)
                elif testtype_mode == 'o':
                    qs_temp = qs.none()
                    for testtype in testtype_list.split(';'):
                        if len(testtype) > 1:
                            match_mode = testtype[0]
                            match_content = testtype[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(testtype__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(testtype__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(testtype__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(testtype__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(testtype__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(testtype__icontains=match_content)
                    qs = qs_temp.distinct()
                    
            tm_m = re.search(r"tm_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if tm_m is not None:
                tm_mode = tm_m.group(1)
                tm_list = tm_m.group(2)
                if tm_mode == 'a':
                    for tm in tm_list.split(';'):
                        if len(tm) > 1:
                            match_mode = tm[0]
                            match_content = tm[1:]
                            if match_mode == 'e':
                                qs = qs.filter(testmodules__name__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(testmodules__name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(testmodules__name__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(testmodules__name__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(testmodules__name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(testmodules__name__icontains=match_content)

                elif tm_mode == 'o':
                    qs_temp = qs.none()
                    for tm in tm_list.split(';'):
                        if len(tm) > 1:
                            match_mode = tm[0]
                            match_content = tm[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(testmodules__name__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(testmodules__name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(testmodules__name__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(testmodules__name__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(testmodules__name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(testmodules__name__icontains=match_content)
                    qs = qs_temp.distinct()

            author_m = re.search(r"author_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if author_m is not None:
                author_mode = author_m.group(1)
                author_list = author_m.group(2)
                if author_mode == 'a':
                    for author in author_list.split(';'):
                        if len(author) > 1:
                            match_mode = author[0]
                            match_content = author[1:]
                            if match_mode == 'e':
                                qs = qs.filter(author__name__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(author__name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(author__name__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(author__name__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(author__name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(author__name__icontains=match_content)
                elif author_mode == 'o':
                    qs_temp = qs.none()
                    for author in author_list.split(';'):
                        if len(author) > 1:
                            match_mode = author[0]
                            match_content = author[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(author__name__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(author__name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(author__name__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(author__name__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(author__name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(author__name__icontains=match_content)
                    qs = qs_temp.distinct()

            tg_m = re.search(r"tg_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if tg_m is not None:
                tg_mode = tg_m.group(1)
                tg_list = tg_m.group(2)
                if tg_mode == 'a':
                    for tg in tg_list.split(';'):
                        if len(tg) > 1:
                            match_mode = tg[0]
                            match_content = tg[1:]
                            if match_mode == 'e':
                                qs = qs.filter(tags__name__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(tags__name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(tags__name__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(tags__name__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(tags__name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(tags__name__icontains=match_content)
                elif tg_mode == 'o':
                    qs_temp = qs.none()
                    for tg in tg_list.split(';'):
                        if len(tg) > 1:
                            match_mode = tg[0]
                            match_content = tg[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(tags__name__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(tags__name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(tags__name__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(tags__name__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(tags__name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(tags__name__icontains=match_content)
                    qs = qs_temp.distinct()
                    
            ms_m = re.search(r"ms_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if ms_m is not None:
                ms_mode = ms_m.group(1)
                ms_list = ms_m.group(2)
                if ms_mode == 'a':
                    for ms in ms_list.split(';'):
                        if len(ms) > 1:
                            match_mode = ms[0]
                            match_content = ms[1:]
                            if match_mode == 'e':
                                qs = qs.filter(mst__name__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(mst__name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(mst__name__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(mst__name__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(mst__name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(mst__name__icontains=match_content)
                elif ms_mode == 'o':
                    qs_temp = qs.none()
                    for ms in ms_list.split(';'):
                        if len(ms) > 1:
                            match_mode = ms[0]
                            match_content = ms[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(mst__name__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(mst__name__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(mst__name__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(mst__name__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(mst__name__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(mst__name__icontains=match_content)
                    qs = qs_temp.distinct()

            duration_m = re.search(r"duration_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if duration_m is not None:
                duration_mode = duration_m.group(1)
                duration_list = duration_m.group(2)
                if duration_mode == 'a':
                    for duration in duration_list.split(';'):
                        if len(duration) > 1:
                            match_mode = duration[0]
                            match_content = duration[1:]
                            if match_mode == 'e':
                                qs = qs.filter(duration_seconds__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(duration_seconds__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(duration_seconds__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(duration_seconds__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(duration_seconds__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(duration_seconds__icontains=match_content)
                elif duration_mode == 'o':
                    qs_temp = qs.none()
                    for duration in duration_list.split(';'):
                        if len(duration) > 1:
                            match_mode = duration[0]
                            match_content = duration[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(duration_seconds__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(duration_seconds__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(duration_seconds__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(duration_seconds__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(duration_seconds__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(duration_seconds__icontains=match_content)
                    qs = qs_temp.distinct()

            defects_m = re.search(r"defects_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if defects_m is not None:
                defects_mode = defects_m.group(1)
                defects_list = defects_m.group(2)
                if defects_mode == 'a':
                    for defects in defects_list.split(';'):
                        if len(defects) > 1:
                            match_mode = defects[0]
                            match_content = defects[1:]
                            if match_mode == 'e':
                                qs = qs.filter(defects__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(defects__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(defects__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(defects__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(defects__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(defects__icontains=match_content)
                elif defects_mode == 'o':
                    qs_temp = qs.none()
                    for defects in defects_list.split(';'):
                        if len(defects) > 1:
                            match_mode = defects[0]
                            match_content = defects[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(defects__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(defects__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(defects__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(defects__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(defects__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(defects__icontains=match_content)
                    qs = qs_temp.distinct()

            timeout_m = re.search(r"timeout_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if timeout_m is not None:
                timeout_mode = timeout_m.group(1)
                timeout_list = timeout_m.group(2)
                if timeout_mode == 'a':
                    for timeout in timeout_list.split(';'):
                        if len(timeout) > 1:
                            match_mode = timeout[0]
                            match_content = timeout[1:]
                            if match_mode == 'e':
                                qs = qs.filter(timeout_seconds__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(timeout_seconds__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(timeout_seconds__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(timeout_seconds__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(timeout_seconds__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(timeout_seconds__icontains=match_content)
                elif timeout_mode == 'o':
                    qs_temp = qs.none()
                    for timeout in timeout_list.split(';'):
                        if len(timeout) > 1:
                            match_mode = timeout[0]
                            match_content = timeout[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(timeout_seconds__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(timeout_seconds__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(timeout_seconds__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(timeout_seconds__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(timeout_seconds__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(timeout_seconds__icontains=match_content)
                    qs = qs_temp.distinct()

            priority_m = re.search(r"priority_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if priority_m is not None:
                priority_mode = priority_m.group(1)
                priority_list = priority_m.group(2)
                if priority_mode == 'a':
                    for priority in priority_list.split(';'):
                        if len(priority) > 1:
                            match_mode = priority[0]
                            match_content = priority[1:]
                            if match_mode == 'e':
                                qs = qs.filter(priority__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(priority__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(priority__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(priority__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(priority__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(priority__icontains=match_content)
                elif priority_mode == 'o':
                    qs_temp = qs.none()
                    for priority in priority_list.split(';'):
                        if len(priority) > 1:
                            match_mode = priority[0]
                            match_content = priority[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(priority__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(priority__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(priority__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(priority__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(priority__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(priority__icontains=match_content)
                    qs = qs_temp.distinct()

            mrq_number_m = re.search(r"mrq_number_(\w+)=([\w;\s\-\_\,\.\|\@]+)", q_str)
            if mrq_number_m is not None:
                mrq_number_mode = mrq_number_m.group(1)
                mrq_number_list = mrq_number_m.group(2)
                print(mrq_number_mode, mrq_number_list)
                if mrq_number_mode == 'a':
                    for mrq_number in mrq_number_list.split(';'):
                        if len(mrq_number) > 1:
                            match_mode = mrq_number[0]
                            match_content = mrq_number[1:]
                            if match_mode == 'e':
                                print("inside mrq e mode")
                                qs = qs.filter(mrq_number__iexact=match_content)
                            elif match_mode == 's':
                                qs = qs.filter(mrq_number__istartswith=match_content)
                            elif match_mode == 'c':
                                qs = qs.filter(mrq_number__icontains=match_content)
                            elif match_mode == 'r':
                                qs = qs.exclude(mrq_number__iexact=match_content)
                            elif match_mode == 'd':
                                qs = qs.exclude(mrq_number__istartswith=match_content)
                            elif match_mode == 'v':
                                qs = qs.exclude(mrq_number__icontains=match_content)
                elif mrq_number_mode == 'o':
                    qs_temp = qs.none()
                    for mrq_number in mrq_number_list.split(';'):
                        if len(mrq_number) > 1:
                            match_mode = mrq_number[0]
                            match_content = mrq_number[1:]
                            if match_mode == 'e':
                                qs_temp = qs_temp | qs.filter(mrq_number__iexact=match_content)
                            elif match_mode == 's':
                                qs_temp = qs_temp | qs.filter(mrq_number__istartswith=match_content)
                            elif match_mode == 'c':
                                qs_temp = qs_temp | qs.filter(mrq_number__icontains=match_content)
                            elif match_mode == 'r':
                                qs_temp = qs_temp | qs.exclude(mrq_number__iexact=match_content)
                            elif match_mode == 'd':
                                qs_temp = qs_temp |  qs.exclude(mrq_number__istartswith=match_content)
                            elif match_mode == 'v':
                                qs_temp = qs_temp |  qs.exclude(mrq_number__icontains=match_content)
                    qs = qs_temp.distinct()

            qs = self.q_str_filter(q_str,'objective',qs)
            qs = self.q_str_filter(q_str,'initial_configuration',qs)
            qs = self.q_str_filter(q_str,'procedure',qs)
            qs = self.q_str_filter(q_str,'expected_results',qs)
            qs = self.q_str_filter(q_str,'environment',qs)
            qs = self.q_str_filter(q_str,'test_plan_section',qs)
            
        return qs            
    
    def formfield_for_manytomany(self, db_field, request, **kwargs):
        if db_field.name == "tags":
            if request.user.is_superuser:
                kwargs["queryset"] = Tag.objects.all().order_by('name')
            else:
                #kwargs["queryset"] = Tag.objects.filter(name__contains=keyword)
                kwargs["queryset"] = TagManager.objects.get(user=request.user.id).tags.all().order_by('name')
        return super(TestCaseAdmin, self).formfield_for_manytomany(db_field, request, **kwargs)

    def getVarForms(self, request, num_var):
        global VAR_EMPTY
        global CURR_VAR_ID
        CURR_VAR_ID = 0
        VariableFormSet = formset_factory(VariableForm, formset=BaseVariableFormSet, extra=num_var)
        if request.method == 'POST':
            varforms = VariableFormSet(request.POST)
            if varforms.is_valid():
                tc = TestCase.objects.get(pk=TC_ID)
                script = TestScript.objects.get(name=tc.script_name.strip())
                variables = script.variables.split(',')
                vArray = []
                for i in range(num_var):
                    if varforms[i].is_valid():
                        vname = variables[i]
                        vval = varforms[i].cleaned_data['value']
                        vArray.append(Variable.objects.get_or_create(name=vname.strip(), value=vval.strip())[0])
                tc.variables = vArray
                tc.save()
            else:
                VAR_EMPTY = True
        else:
            varforms = VariableFormSet()
        CURR_VAR_ID = 0
        return varforms

    
    def change_view(self, request, object_id, form_url='', extra_context=None):
        global TC_ID
        global fieldsets
        extra_context = extra_context or {}
##        temp = request.get_full_path().split('/')
##        
##        TC_ID = temp[len(temp)-2]
##        to_del_view = ''
##        try:
##            #tc = TestCase.objects.get(pk=TC_ID)
##            tc = self.get_object(request, unquote(object_id))
##            tc_vars = tc.variables.all()
##            print tc_vars
##            if tc.exec_harness.exec_harness == 'thot' and tc.script_type.interpreter == 'tcl':
##                script = TestScript.objects.get(name=tc.script_name.strip())
##                if script.variables is None:
##                    scr_vars = []
##                else:
##                    scr_vars = script.variables.split(',')
##                num_var = len(scr_vars)
##                to_del = []
##                for tc_var in tc_vars:
##                    if tc_var.name not in scr_vars:
##                        to_del.append(tc_var.name)
##                to_del_view = ''
##                if len(to_del) > 0:
##                    to_del_view = '<div><font size="5" color="red"><p><b>Variables to delete:</b></p>'
##                    for var in to_del:
##                        to_del_view += '<p>' + var + '</p>'
##                    to_del_view += '</font></div>'
##            else:
##                num_var = len(tc_vars)
##            extra_context['varforms'] = self.getVarForms(request, num_var)
##            extra_context['to_del_vars'] = to_del_view
##        except:
##            extra_context['varforms'] = ''
##            extra_context['to_del_vars'] = to_del_view

        

        result = super(TestCaseAdmin, self).change_view(request, object_id, form_url, extra_context=extra_context)
        ref = request.META.get('HTTP_REFERER', '')
        if ref.find('?') != -1:
            # We've got a query string, set the session value
            request.session['filtered'] =  ref
        if request.POST.has_key('_save'):
            try:
                if request.session['filtered'] is not None:
                    result['Location'] = request.session['filtered']
                    request.session['filtered'] = None
            except:
                pass
            
        return result
        #return super(TestCaseAdmin, self).change_view(request, object_id, form_url, extra_context=extra_context)

    def changelist_view(self, request, extra_context=None):
        extra_context = extra_context or {}
        msg = ''
        
        if not request.user.is_superuser:
            if len(TagManager.objects.get(user=request.user.id).tags.all()) == 0:
                msg = '<p>You have not selected any tags. Please go to \"<a href=\"/stadmin/auth/user/\">My Account</a>\" and select your tags.</p>'
        extra_context['mymsg'] = msg

        return super(TestCaseAdmin, self).changelist_view(request, extra_context=extra_context)
    
    def get_actions(self, request):
        actions = super(TestCaseAdmin, self).get_actions(request)
        del actions['delete_selected']
        return actions
    
    
    #exclude = ('tags',)
    #inlines = [variablesInline]
    filter_horizontal = ['tags', 'testmodules', 'testsets', 'dependencies']
    list_filter = ['modified', 'testsets', 'testmodules', TagFilter, 'state', 'mst', ErrorFilter]
##    class Media:
##        js = ['/static/admin/js/SelectBox.js', '/static/admin/js/SelectFilter2.js']
    search_fields = ['name', 'script_name']
    ordering = ['-modified']
    list_display = ['colored_name', 'modified', 'state']
    actions = [tc_to_text, mass_edit_tc, tc_to_csv,tc_to_xlsx, duplicate_tc]
    form = MyTestCaseAdmin

class TestCaseInline(admin.StackedInline):
    model = TestCase.tags.through
    extra = 0

class TagAdmin(admin.ModelAdmin):

    def queryset(self, request):
        qs = super(TagAdmin, self).queryset(request)
        if request.user.is_superuser:
            return qs.all()
        else:
            tagM = TagManager.objects.get(user=request.user.id)
            return tagM.tags.all()
    
    def get_list_display_links(self, request, list_display):
        if request.user.is_superuser:
            return super(TagAdmin, self).get_list_display_links(request, list_display)
        else:
            return (None, )
    
    def has_add_permission(self, request):
        if request.user.is_superuser:
            return True
        else:
            return False
        
    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        else:
            return False

    def toTcLink(self, obj):
        return '<a href="/stadmin/stapp/testcase/?tag=%s"><b>Browse test cases</b></a>' % (obj.name)
    toTcLink.allow_tags = True
    toTcLink.short_description = 'Link to test cases'

    def get_actions(self, request):
        actions = super(TagAdmin, self).get_actions(request)
        if not request.user.is_superuser:
            del actions['delete_selected']
        return actions

    search_fields = ['name']
    list_display = ['name', 'toTcLink',]

class TestSetAdmin(admin.ModelAdmin):
    def get_list_display_links(self, request, list_display):
        if request.user.is_superuser:
            return super(TestSetAdmin, self).get_list_display_links(request, list_display)
        else:
            return (None, )
    
    def has_add_permission(self, request):
        if request.user.is_superuser:
            return True
        else:
            return False
        
    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        else:
            return False

    def toTcLink(self, obj):
        return '<a href="/stadmin/stapp/testcase/?testsets__id__exact=%s"><b>Browse test cases</b></a>' % (obj.id)
    toTcLink.allow_tags = True
    toTcLink.short_description = 'Link to test cases'

    def get_actions(self, request):
        actions = super(TestSetAdmin, self).get_actions(request)
        if not request.user.is_superuser:
            del actions['delete_selected']
        return actions

    search_fields = ['name']
    list_display = ['name', 'toTcLink',]

class TestModuleAdmin(admin.ModelAdmin):
    def get_list_display_links(self, request, list_display):
        if request.user.is_superuser:
            return super(TestModuleAdmin, self).get_list_display_links(request, list_display)
        else:
            return (None, )
    
    def has_add_permission(self, request):
        if request.user.is_superuser:
            return True
        else:
            return False
        
    def has_delete_permission(self, request, obj=None):
        if request.user.is_superuser:
            return True
        else:
            return False

    def toTcLink(self, obj):
        return '<a href="/stadmin/stapp/testcase/?testmodules__name__exact=%s"><b>Browse test cases</b></a>' % (obj.name)
    toTcLink.allow_tags = True
    toTcLink.short_description = 'Link to test cases'

    def get_actions(self, request):
        actions = super(TestModuleAdmin, self).get_actions(request)
        if not request.user.is_superuser:
            del actions['delete_selected']
        return actions

    search_fields = ['name']
    list_display = ['name', 'STC_Name', 'min_stc_version', 'rerun_on_not_pass', 'threshold_for_rerun_pass', 'module_generation','toTcLink',]

class TagManagerInline(admin.StackedInline):
    model = TagManager
    extra = 0
    fields = ['tags']
    filter_horizontal = ['tags']

    def formfield_for_manytomany(self, db_field, request, **kwargs):
        if db_field.name == "tags":
            kwargs["queryset"] = Tag.objects.all().order_by('name')
        return super(TagManagerInline, self).formfield_for_manytomany(db_field, request, **kwargs)
    
class MyUserAdmin(UserAdmin):
    list_filter = []
        
    list_display = ('username', 'email')

    inlines = [TagManagerInline]
    
    readonly_fields = ['last_login', 'date_joined']

    def get_fieldsets(self, request, obj=None):
        fieldsets = super(MyUserAdmin, self).get_fieldsets(request, obj)
        if not request.user.is_superuser:
            fieldsets = [
                (None, {'fields': ('username', 'password')}),
                ('Personal info',{'fields': ('first_name', 'last_name','email')}),
                ('Permissions', {'fields': ('groups',)}),
                ('Important dates', {'fields': ('last_login', 'date_joined')}),
                ]
        return fieldsets
    
    def queryset(self, request):
        qs = super(MyUserAdmin, self).queryset(request)
        if request.user.is_superuser:
            return qs.all()
        else:
            return qs.filter(username=request.user.username)

    def save_model(self, request, obj, form, change):
        if not request.user.is_superuser:
            obj.is_active = request.user.is_active
            obj.is_staff = request.user.is_staff
        obj.save()    

class DeprecatedVersionAdmin(admin.ModelAdmin):
    def queryset(self, request):
        qs = super(DeprecatedVersionAdmin, self).queryset(request).filter(state=2)
        return qs
    
    def has_add_permission(self, request):
        return False
    
        
    def has_delete_permission(self, request, obj=None):
        return False
    
    readonly_fields = ('name', 'script_name', 'dependencies','variables', 'script_misc_files',
                       'testmodules', 'tags', 'author', 'state',
                       'p4_script_version', 'script_type', 'duration_seconds','defects',
                       'timeout_seconds', 'mrq_number', 'testsets', 'testtype',
                       'min_stc_version', 'exec_harness', 'mst', 'priority', 'objective','procedure','test_plan_section','environment', 
                       'expected_results','initial_configuration', 'previous_version')
    
    ordering = ['-modified']
    list_display = ['name', 'modified']
    search_fields = ['name', 'script_name']
    actions = None
    exclude = ['misc_file_location',]
    
class DeprecatedVersion(TestCase):
    class Meta:
        proxy = True
    

class PromotedTestCaseAdmin(admin.ModelAdmin):
    def queryset(self, request):
        qs = super(PromotedTestCaseAdmin, self).queryset(request).filter(state=0)
        return qs
    
    def has_add_permission(self, request):
        return False
    
        
    def has_delete_permission(self, request, obj=None):
        return False
    
    readonly_fields = ('name', 'script_name', 'dependencies','variables', 'script_misc_files',
                       'testmodules', 'tags', 'author', 'state',
                       'p4_script_version', 'script_type', 'duration_seconds','defects',
                       'timeout_seconds', 'mrq_number', 'testsets', 'testtype',
                       'min_stc_version', 'exec_harness', 'mst', 'priority', 'objective','procedure','test_plan_section','environment', 
                       'expected_results','initial_configuration', 'previous_version')
    
    ordering = ['-modified']
    list_display = ['name', 'modified']
    search_fields = ['name', 'script_name']
    actions = None
    exclude = ['misc_file_location',]
    
class PromotedTestCase(TestCase):
    class Meta:
        proxy = True


class VersionInline(admin.StackedInline):
    model = reversion.models.Version
    extra = 0
    exclude = ('object_id', 'object_id_int',)
    readonly_fields = ['content_type','format','serialized_data','object_repr','type',]


class MyRevision(admin.ModelAdmin):    
    list_display = ['test_cases', 'user', 'comment','date_created', 'rev_link', 'undo_link']
    exclude = ('manager_slug',)
    readonly_fields = ['user','comment']
    inlines = [VersionInline]
    actions = [revert_selected,]

    def queryset(self, request):
        qs = super(MyRevision, self).queryset(request)
        if request.user.is_superuser:
            return qs.all()
        else:
            return qs.filter(user=request.user.pk)

    def __init__(self, *args, **kwargs):
        super(MyRevision, self).__init__(*args, **kwargs)
        #self.list_display_links = (None, )

    def test_cases(self, obj):
        return obj.__unicode__()[:100]

    def rev_link(self, obj):
        vers = reversion.models.Version.objects.filter(revision_id=obj.id)
        for ver in vers:
            if ver.type == 2:
                return ''
        return '<a href="/stapp/revision/%s/"><b>Revert</b></a>' % (obj.id)
    rev_link.allow_tags = True
    rev_link.short_description = ''

    def undo_link(self, obj):
        return '<a href="/stapp/undo/%s/"><b>Undo</b></a>' % (obj.id)
    undo_link.allow_tags = True
    undo_link.short_description = ''

    def get_actions(self, request):
        actions = super(MyRevision, self).get_actions(request)
        del actions['delete_selected']
        return actions


class MyVersion(admin.ModelAdmin):
    exclude = ('object_id', 'object_id_int',)
    readonly_fields = ['content_type','format','serialized_data','object_repr','type',]


class DBFlagAdmin(admin.ModelAdmin):
	fieldsets = [
		(None, {'fields':['name']}),
		('Flag Value',{'fields':['value']}),
	]
	def get_actions(self, request):
		actions = super(DBFlagAdmin, self).get_actions(request)
		del actions['delete_selected']
		return actions
	
	def get_readonly_fields(self,request,obj=None):
		if obj and obj.name == "adminlocked":
			return ('name',)
		else:
			return ('name','value')
	
	
	def change_view(self, request, object_id, form_url='', extra_context=None):
		extra_context = extra_context or {}
		extra_context['show_save_and_add_another'] = False
		extra_context['show_save_and_continue'] = False
		return super(DBFlagAdmin, self).change_view(request, object_id, form_url, extra_context=extra_context)
	
	
	def has_delete_permission(self, request, obj=None):
		return False
	
	def has_add_permission(self, request):
		return False
	
	list_display = ['name', 'tf_value', 'last_update',]


admin.site.unregister(User)
admin.site.register(User,MyUserAdmin)
admin.site.unregister(Site)

#admin.site.register(TestCaseAuthor, TestCaseAuthorAdmin)
admin.site.register(TestCase, TestCaseAdmin)
admin.site.register(Tag,TagAdmin)
admin.site.register(Variable)
admin.site.register(TestModule, TestModuleAdmin)
admin.site.register(DBFlag, DBFlagAdmin)
admin.site.register(TestSet, TestSetAdmin)
admin.site.register(MiscFileLocation)
admin.site.register(Dependency)
admin.site.register(ScriptMiscFile)
admin.site.register(reversion.models.Revision, MyRevision)
