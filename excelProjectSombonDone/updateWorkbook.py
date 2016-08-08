import os
import csv
from openpyxl import load_workbook
import subprocess
import sys
import json
files = []
path = os.path.dirname(os.path.abspath(__file__))
for i in os.listdir(path):
    files.append(i)
#determine the spreadsheet name and row to start inserting new data
updateFiles = []
for filename in files:
	if "kw_integration_win_ui_diff" in filename:
		updateFiles.append({filename:"New UI"})
	if "kw_integration_Win_BLL_baseline_diff" in filename:
		updateFiles.append({filename:"New BLL"})
	if "integration_x86_baseline_uninit" in filename:
		updateFiles.append({filename:"UNINIT IL"})
	if "integration_x86_baseline_mem_leak_" in filename:
		updateFiles.append({filename:"IL Leaks"})
	if "integration_x86_baseline_backlog_" in filename:
		updateFiles.append({filename:"BKLG IL"})
	if "integration_win_ui_mem_leak_" in filename:
		updateFiles.append({filename:"UI Leaks"})
	if "integration_win_ui_backlog_" in filename:
		updateFiles.append({filename:"BKLG UI"})
	if "integration_Win_BLL_baseline_uninit" in filename:
		updateFiles.append({filename:"UNINIT BLL"})
	if "integration_Win_BLL_baseline_mem_leak" in filename:
		updateFiles.append({filename:"BLL Leaks"})
	if "integration_Win_BLL_baseline_backlog" in filename:
		updateFiles.append({filename:"BKLG BLL"})
	if "kw_integration_x86_baseline_diff" in filename:
		updateFiles.append({filename:"New IL"})
#get the new data, and append each row into a list
def updateCSV(origFile, updateFiles):
	#input validation
	if origFile not in files:
		print "Error: file "+origFile+" could not be found"
		return
	if len(updateFiles) < 1:
		print "To update the xlsm workbook, please add CSV files in the current directory"
		return
	#start updating the sheets
	for updateFile in updateFiles:
		for key in updateFile:
			filename = key
			sheetname = updateFile[key]
		newData = []
		with open(filename, 'rb') as f:
			reader = csv.reader(f,delimiter=";")
			for row in reader:
				newData.append(row)
		wb = load_workbook(filename=origFile,keep_vba=True )
		sheet = wb.get_sheet_by_name(sheetname)
		i = 0
		startUpdate =0
		for row in range(1, sheet.max_row+len(newData)):
			if i > len(newData)-1:
				if sheet["A"+str(row)].value == None:
					print "finished updating sheet "+sheetname
					break
				else:
					sheet["A"+str(row)].value = ""
					sheet["B"+str(row)].value = ""
					sheet["C"+str(row)].value = ""
			else:
				if startUpdate:
					sheet["A"+str(row)].value = int(newData[i][0])
					sheet["B"+str(row)].value = newData[i][1]
					sheet["C"+str(row)].value = newData[i][2]
					i = i+1
				#only start changing lines after the Issue # row. 
				if sheet["A"+str(row)].value == "Issue #":
					startUpdate = 1
		wb.save(origFile)
def updateIPS_PGA(origFile):
	#connect to klocwork
	user = "spirent"
	host = "http://spccalengcov01.spirentcom.com:8090/review/api"

	#get all successfull build dates for a project
	builds = subprocess.check_output(["curl","-s", "--data", "action=builds&user="+user+"&project=integration_Win_BLL_baseline", host])
	buildslist = builds.split("\n")
	dates = []
	for i in buildslist:
		try:
			dates.append(json.loads(i)['name'].split("_")[1])
		except:
			continue
	dates.reverse()
	IpsViews = ["IPS Benchmarking","IPS Custom","IPS DCB"]
	PGAViews = ["PGA Capture", "PGA IEEE 802.11", "PGA L2L3", "PGA Learning"]
	dataLocation = ["A","B","C", "D","F", "G", "H", "I"]
	sheetname = "IPS_PGA_Win_BLL"
	print "Adding data to sheet: "+sheetname
	if origFile not in files:
		print "Error: file "+origFile+" could not be found"
		return
	wb = load_workbook(filename=origFile,keep_vba=True )
	sheet = wb.get_sheet_by_name(sheetname)
	startUpdate = 0
	p = 0
	for row in range(1,sheet.max_row+len(dates)):
		#find where the dates start so you know where  to insert the values
		#start inserting the data
		if startUpdate == 1:
			try:
				date = dates[p]
			except:
				continue
			p = p+1
			sheetVals = []
			sheetVals.append(date)
			#curl to get IPS and PGS values
			for IpsView in IpsViews:
				results = subprocess.check_output(["curl","-s", "--data","action=search&user="+user+"&project=integration_Win_BLL_baseline&query=status:Analyze,Fix build:'integration_"+date+"'&view="+IpsView,host])
				sheetVals.append(len(results.split("\n"))-1)
			for PGAView in PGAViews:
				results = subprocess.check_output(["curl","-s", "--data","action=search&user="+user+"&project=integration_Win_BLL_baseline&query=status:Analyze,Fix build:'integration_"+date+"'&view="+PGAView,host])
				sheetVals.append(len(results.split("\n"))-1)
			for i in range(0,len(sheetVals)):
				sheet[dataLocation[i]+str(row)].value = sheetVals[i]
		if row > 1:
			if sheet["A"+str(row-1)].value == "Date":
				startUpdate = 1
	wb.save(origFile)
if __name__ == "__main__":
	updateCSV(sys.argv[1],updateFiles) 
	updateIPS_PGA(sys.argv[1])
